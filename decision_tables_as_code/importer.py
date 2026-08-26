from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Iterable, Mapping

import yaml


class SpreadsheetImportError(ValueError):
    """Raised when a spreadsheet cannot be converted deterministically."""


def load_import_config(path: str | Path) -> Mapping[str, Any]:
    source = Path(path)
    text = source.read_text(encoding="utf-8")
    if source.suffix.lower() == ".json":
        raw = json.loads(text)
    elif source.suffix.lower() in {".yaml", ".yml"}:
        raw = yaml.safe_load(text)
    else:
        raise SpreadsheetImportError("Import config must be YAML or JSON")
    if not isinstance(raw, Mapping):
        raise SpreadsheetImportError("Import config must contain an object at the root")
    return raw


def import_spreadsheet(source: str | Path, config: Mapping[str, Any]) -> dict[str, Any]:
    path = Path(source)
    table_config = _mapping(config.get("table"), "table")
    columns = _mapping(config.get("columns"), "columns")
    input_config = _mapping(columns.get("inputs"), "columns.inputs")
    output_config = _mapping(columns.get("outputs"), "columns.outputs")
    if not input_config:
        raise SpreadsheetImportError("columns.inputs must define at least one input")
    if not output_config:
        raise SpreadsheetImportError("columns.outputs must define at least one output")

    blank_condition = str(config.get("blank_condition", "wildcard")).lower()
    if blank_condition not in {"wildcard", "error", "empty"}:
        raise SpreadsheetImportError("blank_condition must be wildcard, error, or empty")

    rows, headers = _read_rows(path, sheet=config.get("sheet"))
    _validate_columns(headers, columns, input_config, output_config)

    rules: list[dict[str, Any]] = []
    for row_number, row in rows:
        if _row_is_blank(row):
            continue
        rules.append(
            _build_rule(
                row=row,
                row_number=row_number,
                columns=columns,
                input_config=input_config,
                output_config=output_config,
                blank_condition=blank_condition,
            )
        )

    table_id = table_config.get("id")
    if not isinstance(table_id, str) or not table_id.strip():
        raise SpreadsheetImportError("table.id must be a non-empty string")

    result: dict[str, Any] = {
        "version": int(table_config.get("version", 1)),
        "id": table_id,
        "name": table_config.get("name", table_id),
        "hit_policy": str(table_config.get("hit_policy", "unique")).lower(),
        "inputs": [_build_input(name, value) for name, value in input_config.items()],
        "outputs": [_build_output(name, value) for name, value in output_config.items()],
        "rules": rules,
    }
    for optional in ("description", "metadata"):
        if optional in table_config:
            result[optional] = table_config[optional]
    return result


def dump_yaml(document: Mapping[str, Any]) -> str:
    return yaml.safe_dump(dict(document), sort_keys=False, allow_unicode=True, width=120)


def parse_condition_cell(value: Any, data_type: str, *, blank_condition: str = "wildcard") -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        if blank_condition == "wildcard":
            return "*"
        if blank_condition == "empty":
            return ""
        raise SpreadsheetImportError("Blank condition cell is not allowed")

    if not isinstance(value, str):
        return _coerce_scalar(value, data_type)

    text = value.strip()
    if text == "*":
        return "*"

    comparisons = (
        (">=", "gte"),
        ("<=", "lte"),
        ("!=", "ne"),
        (">", "gt"),
        ("<", "lt"),
        ("=", "eq"),
    )
    for prefix, operator in comparisons:
        if text.startswith(prefix):
            return {operator: _coerce_scalar(text[len(prefix):].strip(), data_type)}

    function_match = re.fullmatch(r"([A-Za-z_]+)\((.*)\)", text)
    if function_match:
        operator = function_match.group(1).lower()
        body = function_match.group(2)
        if operator in {"in", "not_in"}:
            values = [_coerce_scalar(item.strip(), data_type) for item in _split_args(body)]
            if not values:
                raise SpreadsheetImportError(f"{operator}() requires at least one value")
            return values if operator == "in" else {"not_in": values}
        if operator == "between":
            values = [_coerce_scalar(item.strip(), data_type) for item in _split_args(body)]
            if len(values) != 2:
                raise SpreadsheetImportError("between() requires exactly two values")
            return {"between": values}
        if operator == "regex":
            if data_type != "string":
                raise SpreadsheetImportError("regex() is only valid for string inputs")
            return {"regex": body}
        if operator == "exists":
            return {"exists": _parse_bool(body.strip())}
        raise SpreadsheetImportError(f"Unsupported condition expression {operator}()")

    return _coerce_scalar(text, data_type)


def parse_output_cell(value: Any, data_type: str) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    return _coerce_scalar(value.strip() if isinstance(value, str) else value, data_type)


def _build_rule(
    *,
    row: Mapping[str, Any],
    row_number: int,
    columns: Mapping[str, Any],
    input_config: Mapping[str, Any],
    output_config: Mapping[str, Any],
    blank_condition: str,
) -> dict[str, Any]:
    id_column = columns.get("rule_id")
    if id_column:
        raw_id = row.get(str(id_column))
        if raw_id is None or not str(raw_id).strip():
            raise SpreadsheetImportError(f"Row {row_number}: rule id is blank")
        rule_id = str(raw_id).strip()
    else:
        rule_id = f"row-{row_number:04d}"

    rule: dict[str, Any] = {"id": rule_id, "when": {}, "then": {}}

    priority_column = columns.get("priority")
    if priority_column:
        raw_priority = row.get(str(priority_column))
        if raw_priority not in (None, ""):
            try:
                rule["priority"] = int(raw_priority)
            except (TypeError, ValueError) as exc:
                raise SpreadsheetImportError(f"Row {row_number}: priority must be an integer") from exc

    description_column = columns.get("description")
    if description_column:
        raw_description = row.get(str(description_column))
        if raw_description not in (None, ""):
            rule["description"] = str(raw_description).strip()

    for name, definition_raw in input_config.items():
        definition = _mapping(definition_raw, f"columns.inputs.{name}")
        column = _required_column(definition, f"columns.inputs.{name}")
        data_type = str(definition.get("type", "string")).lower()
        try:
            rule["when"][name] = parse_condition_cell(row.get(column), data_type, blank_condition=blank_condition)
        except SpreadsheetImportError as exc:
            raise SpreadsheetImportError(f"Row {row_number}, input {name!r}: {exc}") from exc

    for name, definition_raw in output_config.items():
        definition = _mapping(definition_raw, f"columns.outputs.{name}")
        column = _required_column(definition, f"columns.outputs.{name}")
        data_type = str(definition.get("type", "string")).lower()
        try:
            rule["then"][name] = parse_output_cell(row.get(column), data_type)
        except SpreadsheetImportError as exc:
            raise SpreadsheetImportError(f"Row {row_number}, output {name!r}: {exc}") from exc
    return rule


def _build_input(name: str, raw: Any) -> dict[str, Any]:
    definition = _mapping(raw, f"columns.inputs.{name}")
    result: dict[str, Any] = {"name": name, "type": str(definition.get("type", "string")).lower()}
    for optional in ("description", "domain"):
        if optional in definition:
            result[optional] = definition[optional]
    return result


def _build_output(name: str, raw: Any) -> dict[str, Any]:
    definition = _mapping(raw, f"columns.outputs.{name}")
    result: dict[str, Any] = {"name": name, "type": str(definition.get("type", "string")).lower()}
    if "description" in definition:
        result["description"] = definition["description"]
    return result


def _read_rows(path: Path, *, sheet: Any = None) -> tuple[list[tuple[int, dict[str, Any]]], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = [str(item) for item in (reader.fieldnames or [])]
            rows = [(index, dict(row)) for index, row in enumerate(reader, start=2)]
        return rows, headers

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SpreadsheetImportError("XLSX import requires `pip install decision-tables-as-code[excel]`") from exc
        workbook = load_workbook(path, read_only=True, data_only=False)
        if sheet is None:
            worksheet = workbook[workbook.sheetnames[0]]
        else:
            if str(sheet) not in workbook.sheetnames:
                raise SpreadsheetImportError(f"Worksheet {sheet!r} does not exist")
            worksheet = workbook[str(sheet)]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise SpreadsheetImportError("Spreadsheet is empty") from exc
        headers = ["" if value is None else str(value).strip() for value in header_row]
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(iterator, start=2):
            row = dict(zip(headers, values))
            formulas = [header for header, value in row.items() if isinstance(value, str) and value.startswith("=")]
            if formulas:
                raise SpreadsheetImportError(
                    f"Row {row_number}: formulas are not evaluated; replace formula cells with values: {', '.join(formulas)}"
                )
            rows.append((row_number, row))
        return rows, headers

    raise SpreadsheetImportError("Spreadsheet must be .csv or .xlsx")


def _validate_columns(
    headers: Iterable[str],
    columns: Mapping[str, Any],
    input_config: Mapping[str, Any],
    output_config: Mapping[str, Any],
) -> None:
    header_set = set(headers)
    required: list[tuple[str, str]] = []
    for key in ("rule_id", "priority", "description"):
        value = columns.get(key)
        if value:
            required.append((key, str(value)))
    for name, raw in input_config.items():
        required.append((f"input {name}", _required_column(_mapping(raw, f"columns.inputs.{name}"), f"columns.inputs.{name}")))
    for name, raw in output_config.items():
        required.append((f"output {name}", _required_column(_mapping(raw, f"columns.outputs.{name}"), f"columns.outputs.{name}")))

    missing = [f"{role}: {column!r}" for role, column in required if column not in header_set]
    if missing:
        raise SpreadsheetImportError("Missing spreadsheet columns: " + "; ".join(missing))


def _coerce_scalar(value: Any, data_type: str) -> Any:
    if data_type in {"string", "date", "any"}:
        return str(value).strip() if not isinstance(value, str) else value.strip()
    if data_type == "integer":
        if isinstance(value, bool):
            raise SpreadsheetImportError("boolean cannot be converted to integer")
        try:
            number = float(value)
            if not number.is_integer():
                raise ValueError
            return int(number)
        except (TypeError, ValueError) as exc:
            raise SpreadsheetImportError(f"Expected integer, got {value!r}") from exc
    if data_type == "number":
        if isinstance(value, bool):
            raise SpreadsheetImportError("boolean cannot be converted to number")
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise SpreadsheetImportError(f"Expected number, got {value!r}") from exc
        return int(number) if number.is_integer() else number
    if data_type == "boolean":
        if isinstance(value, bool):
            return value
        return _parse_bool(str(value).strip())
    raise SpreadsheetImportError(f"Unsupported configured type {data_type!r}")


def _parse_bool(value: str) -> bool:
    normalized = value.strip().lower()
    if normalized in {"true", "yes", "1"}:
        return True
    if normalized in {"false", "no", "0"}:
        return False
    raise SpreadsheetImportError(f"Expected boolean, got {value!r}")


def _split_args(value: str) -> list[str]:
    return [item for item in (part.strip() for part in value.split(",")) if item]


def _mapping(value: Any, path: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise SpreadsheetImportError(f"{path} must be an object")
    return value


def _required_column(definition: Mapping[str, Any], path: str) -> str:
    value = definition.get("column")
    if not isinstance(value, str) or not value.strip():
        raise SpreadsheetImportError(f"{path}.column must be a non-empty string")
    return value


def _row_is_blank(row: Mapping[str, Any]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in row.values())
